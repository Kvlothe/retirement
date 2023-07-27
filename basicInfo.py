from datetime import datetime
from tkinter import ttk
import openpyxl
import tkinter as tk


# def main_menu_window():
#     global main_window
#     main_window = tk.Tk()
#     main_window.geometry("400x400")
#     # basic = tk.Button(main_window, text="Basic info", command=basicInfo.basic_info_window)
#     basic = tk.Button(main_window, text="Basic info", command=lambda: basicInfo.basic_info_window(wb, ws, username))
#     basic.pack()
#     # current = tk.Button(main_window, text="Current Profile", command=currentProfile.current_profile_window)
#     current = tk.Button(main_window, text="Current Profile", command=lambda: currentProfile.current_profile_window(wb, ws, username))
#     current.pack()
#     social_security = tk.Button(main_window, text='Social Security', command=socialSecurity.security_window)
#     social_security.pack()
#     pension = tk.Button(main_window, text="Pension (Monthly)", command=pension.pension_window)
#     pension.pack()
#     estimated_expenses = tk.Button(main_window, text="Estimated Expenses", command=estimatedExpenses.expenses_window)
#     estimated_expenses.pack()


def update_basic_info(wb, ws, username):
    # global main_window
    current_year = year_entry.get()
    age = age_entry.get()
    retirement_age = retirement_age_combo.get()
    ss_s1 = ss_s1_combo.get()
    ss_s2 = ss_s2_combo.get()
    pension_s1 = pension_s1_combo.get()
    pension_s2 = pension_s2_combo.get()
    other_income_s1 = other_income_s1_combo.get()
    other_income_s2 = other_income_s2_combo.get()

    ws['C5'] = current_year
    ws['C6'] = age
    ws['C7'] = retirement_age
    ws['C8'] = ss_s1
    ws['E8'] = ss_s2
    ws['C9'] = pension_s1
    ws['E9'] = pension_s2
    ws['C10'] = other_income_s1
    ws['E10'] = other_income_s2

    wb.save(f'{username}.xlsx')


def basic_info_window(window, wb, ws, username):
    global main_window, year_entry, age_entry, retirement_age_combo, ss_s1_combo, ss_s2_combo, pension_s1_combo, pension_s2_combo, other_income_s1_combo, other_income_s2_combo
    basic_window = tk.Tk()
    basic_window.geometry("400x500")
    # Here you can add your widgets that display the info from the loaded workbook.

    # Create a Label and an Entry for Current Year
    year_label = tk.Label(basic_window, text="Current Year:")
    year_label.pack()
    year_entry = tk.Entry(basic_window)
    year_entry.pack()

    # Create a Label and Entry for Age
    age_label = tk.Label(basic_window, text="Enter Your age")
    age_label.pack()
    age_entry = tk.Entry(basic_window)
    age_entry.pack()

    # Create a Combobox for Retirement age
    age_to_retire = tk.Label(basic_window, text="Retirement Age")
    age_to_retire.pack()
    retirement_age_combo = ttk.Combobox(basic_window, values=["62", "65", "67", "70"])
    retirement_age_combo.pack()
    # retirement_age_combo.bind('<<ComboboxSelected>>', combo_selection)

    year_take_ss_label = tk.Label(basic_window, text="Age to Take Social Security")
    year_take_ss_label.pack()
    # Create a Combobox for SS
    year_ss_s1 = tk.Label(basic_window, text="Self")
    year_ss_s1.pack()
    ss_s1_combo = ttk.Combobox(basic_window, values=["62", "65", "67", "70"])
    ss_s1_combo.pack()

    # Create a Combobox for SS Spouse
    year_ss_s2 = tk.Label(basic_window, text="Spouse")
    year_ss_s2.pack()
    ss_s2_combo = ttk.Combobox(basic_window, values=["62", "65", "67", "70"])
    ss_s2_combo.pack()

    year_take_pension_label = tk.Label(basic_window, text="Age to take Pension")
    year_take_pension_label.pack()
    # Create a Combobox for pension
    pension_s1 = tk.Label(basic_window, text="Self")
    pension_s1.pack()
    pension_s1_combo = ttk.Combobox(basic_window, values=["62", "65", "lump"])
    pension_s1_combo.pack()

    # Create a Combobox for pension for spouse
    pension_s2 = tk.Label(basic_window, text="Spouse")
    pension_s2.pack()
    pension_s2_combo = ttk.Combobox(basic_window, values=["62", "65", "lump"])
    pension_s2_combo.pack()

    year_other_income_label = tk.Label(basic_window, text="Year to take Other Income  (Annuity, ETC)")
    year_other_income_label.pack()
    # Create a Combobox for other incomes
    other_income_s1 = tk.Label(basic_window, text="Self")
    other_income_s1.pack()
    other_income_s1_combo = ttk.Combobox(basic_window, values=["62", "65", "67", "70"])
    other_income_s1_combo.pack()
    other_income_s2 = tk.Label(basic_window, text="Spouse")
    other_income_s2.pack()
    other_income_s2_combo = ttk.Combobox(basic_window, values=["62", "65", "67", "70"])
    other_income_s2_combo.pack()

    # submit_basic = tk.Button(basic_window, text="Submit", command=update_basic_info)
    submit_basic = tk.Button(basic_window, text="Submit", command=lambda: update_basic_info(wb, ws, username))
    submit_basic.pack()

    # Populate the Entry field with the current year
    current_year = datetime.now().year
    year_entry.insert(0, str(current_year))

    basic_window.mainloop()
