import tkinter as tk
from openpyxl.styles import numbers


def update_pension_info(wb, ws, username):
    self_pension_63 = pension_s1_63_entry.get()
    self_pension_65 = pension_s1_65_entry.get()
    self_pension_lump = pension_s1_lump_entry.get()
    spouse_pension_63 = pension_s2_63_entry.get()
    spouse_pension_65 = pension_s2_65_entry.get()
    spouse_pension_lump = pension_s2_lump_entry.get()

    ws['C22'] = int(self_pension_63)
    ws['D22'] = int(self_pension_65)
    ws['E22'] = int(self_pension_lump)
    ws['C23'] = int(spouse_pension_63)
    ws['D23'] = int(spouse_pension_65)
    ws['E23'] = int(spouse_pension_lump)
    ws['C22'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['D22'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['E22'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C23'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['D23'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['E23'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(f'{username}.xlsx')


def pension_input_window(pension_window, wb, ws, username):
    global pension_s1_63_entry, pension_s1_65_entry, pension_s1_lump_entry, pension_s2_63_entry, pension_s2_65_entry, pension_s2_lump_entry

    pension_window.geometry("200x400")

    pension_s1_63_value = ws['C22'].value if ws['C22'].value else ""
    pension_s1_65_value = ws['D22'].value if ws['D22'].value else ""
    pension_s1_lump_value = ws['E22'].value if ws['E22'].value else ""
    pension_s2_63_value = ws['C23'].value if ws['C23'].value else ""
    pension_s2_65_value = ws['D23'].value if ws['D23'].value else ""
    pension_s2_lump_value = ws['E23'].value if ws['E23'].value else ""

    # Create Label and Entry for Pensions
    pension_self_label = tk.Label(pension_window, text="Self")
    pension_self_label.pack()
    pension_63_label = tk.Label(pension_window, text="63")
    pension_63_label.pack()
    pension_s1_63_entry = tk.Entry(pension_window)
    pension_s1_63_entry.insert(0, pension_s1_63_value)
    pension_s1_63_entry.pack()
    pension_65_label = tk.Label(pension_window, text="65")
    pension_65_label.pack()
    pension_s1_65_entry = tk.Entry(pension_window)
    pension_s1_65_entry.insert(0, pension_s1_65_value)
    pension_s1_65_entry.pack()
    pension_lump_label = tk.Label(pension_window, text="Lump")
    pension_lump_label.pack()
    pension_s1_lump_entry = tk.Entry(pension_window)
    pension_s1_lump_entry.insert(0, pension_s1_lump_value)
    pension_s1_lump_entry.pack()
    pension_spouse_label = tk.Label(pension_window, text="Spouse")
    pension_spouse_label.pack()
    spouse_pension_63_label = tk.Label(pension_window, text="63")
    spouse_pension_63_label.pack()
    pension_s2_63_entry = tk.Entry(pension_window)
    pension_s2_63_entry.insert(0, pension_s2_63_value)
    pension_s2_63_entry.pack()
    spouse_pension_65_label = tk.Label(pension_window, text="65")
    spouse_pension_65_label.pack()
    pension_s2_65_entry = tk.Entry(pension_window)
    pension_s2_65_entry.insert(0, pension_s2_65_value)
    pension_s2_65_entry.pack()
    spouse_pension_lump_label = tk.Label(pension_window, text="Lump")
    spouse_pension_lump_label.pack()
    pension_s2_lump_entry = tk.Entry(pension_window)
    pension_s2_lump_entry.insert(0, pension_s2_lump_value)
    pension_s2_lump_entry.pack()

    submit_basic = tk.Button(pension_window, text="Submit", command=lambda: update_pension_info(wb, ws, username))
    submit_basic.pack()

    pension_window.mainloop()
