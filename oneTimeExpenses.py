import tkinter as tk
from openpyxl.styles import numbers


def update_one_time_info(wb, ws, username):
    move = move_entry.get()
    retirement_party = retirement_party_entry.get()
    initial_big_trip = initial_big_trip_entry.get()
    home_improvement = home_improvement_entry.get()
    gifts = gifts_entry.get()
    car_motor_home = car_motor_home_entry.get()
    annuity = annuity_entry.get()
    tax = tax_for_what_we_use_here_entry.get()
    misc = miscellaneous_entry.get()

    ws['K28'] = int(move)
    ws['K29'] = int(retirement_party)
    ws['K30'] = int(initial_big_trip)
    ws['K31'] = int(home_improvement)
    ws['K32'] = int(gifts)
    ws['K33'] = int(car_motor_home)
    ws['K34'] = int(annuity)
    ws['K35'] = int(tax)
    ws['K36'] = int(misc)
    ws['K28'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K29'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K30'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K31'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K32'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K33'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K34'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K35'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['K36'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(f'{username}.xlsx')


def one_time_window(one_time_expense_window, wb, ws, username):
    global move_entry, retirement_party_entry, initial_big_trip_entry, home_improvement_entry, gifts_entry, car_motor_home_entry, annuity_entry, tax_for_what_we_use_here_entry, miscellaneous_entry

    one_time_expense_window.geometry("400x400")

    move_value = ws['K28'].value if ws['K28'].value else ""
    party_value = ws['K29'].value if ws['K29'].value else ""
    trip_value = ws['K30'].value if ws['K30'].value else ""
    improvement_value = ws['K31'].value if ws['K31'].value else ""
    gifts_value = ws['K32'].value if ws['K32'].value else ""
    car_home_value = ws['K33'].value if ws['K33'].value else ""
    annuity_value = ws['K34'].value if ws['K34'].value else ""
    tax_value = ws['K35'].value if ws['K35'].value else ""
    misc_value = ws['K36'].value if ws['K36'].value else ""

    # Create Labels and Entries for One Time Expense
    move_label = tk.Label(one_time_expense_window, text="Move???")
    move_label.pack()
    move_entry = tk.Entry(one_time_expense_window)
    move_entry.insert(0, move_value)
    move_entry.pack()
    retirement_party_label = tk.Label(one_time_expense_window, text="Retirement Party")
    retirement_party_label.pack()
    retirement_party_entry = tk.Entry(one_time_expense_window)
    retirement_party_entry.insert(0, party_value)
    retirement_party_entry.pack()
    initial_big_trip_label = tk.Label(one_time_expense_window, text="Initial Big Trip")
    initial_big_trip_label.pack()
    initial_big_trip_entry = tk.Entry(one_time_expense_window)
    initial_big_trip_entry.insert(0, trip_value)
    initial_big_trip_entry.pack()
    home_improvement_label = tk.Label(one_time_expense_window, text="Home Improvement")
    home_improvement_label.pack()
    home_improvement_entry = tk.Entry(one_time_expense_window)
    home_improvement_entry.insert(0, improvement_value)
    home_improvement_entry.pack()
    gifts_label = tk.Label(one_time_expense_window, text="Gifts")
    gifts_label.pack()
    gifts_entry = tk.Entry(one_time_expense_window)
    gifts_entry.insert(0, gifts_value)
    gifts_entry.pack()
    car_motor_home_label = tk.Label(one_time_expense_window, text="Car or Motor home")
    car_motor_home_label.pack()
    car_motor_home_entry = tk.Entry(one_time_expense_window)
    car_motor_home_entry.insert(0, car_home_value)
    car_motor_home_entry.pack()
    annuity_label = tk.Label(one_time_expense_window, text="Annuity**", bg="red")
    annuity_label.pack()
    annuity_entry = tk.Entry(one_time_expense_window)
    annuity_entry.insert(0, annuity_value)
    annuity_entry.pack()
    tax_for_what_we_use_here_label = tk.Label(one_time_expense_window, text="Tax for what we use here", bg="red")
    tax_for_what_we_use_here_label.pack()
    tax_for_what_we_use_here_entry = tk.Entry(one_time_expense_window)
    tax_for_what_we_use_here_entry.insert(0, tax_value)
    tax_for_what_we_use_here_entry.pack()
    miscellaneous_label = tk.Label(one_time_expense_window, text="Miscellaneous")
    miscellaneous_label.pack()
    miscellaneous_entry = tk.Entry(one_time_expense_window)
    miscellaneous_entry.insert(0, misc_value)
    miscellaneous_entry.pack()

    submit_basic = tk.Button(one_time_expense_window, text="Submit", command=lambda: update_one_time_info(wb, ws, username))
    submit_basic.pack()

    one_time_expense_window.mainloop()
