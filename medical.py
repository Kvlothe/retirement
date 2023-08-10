import tkinter as tk
from openpyxl.styles import numbers
# from gui import ToolTip


def update_medical_info(wb, ws, username):
    medical_info = medical_expense_entry.get()
    ws['C45'] = int(medical_info)
    ws['C45'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    wb.save(f'{username}.xlsx')


class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip, text=self.text, bg="yellow", relief=tk.SOLID, borderwidth=1)
        label.pack(ipadx=1)

    def leave(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()


def medical_window(medical_input_window, wb, ws, username):
    global medical_expense_entry

    # medical_input_window = tk.Tk()
    medical_input_window.geometry("100x50")

    medical_value = ws['C45'].value if ws['C45'].value else ""

    # Create Label and Entry for Medical
    medical_expense_label = tk.Label(medical_input_window, text="Medical")
    medical_expense_label.pack()

    medical_expense_entry = tk.Entry(medical_input_window)
    medical_expense_entry.insert(0, medical_value)
    medical_expense_entry.pack()

    # Add tooltip to the medical_expense_entry
    ToolTip(medical_expense_entry, "Enter your medical expenses in USD.")

    submit_basic = tk.Button(medical_input_window, text="Submit", command=lambda: update_medical_info(wb, ws, username))
    submit_basic.pack()

    medical_input_window.mainloop()
