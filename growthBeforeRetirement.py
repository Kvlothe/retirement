import tkinter as tk
from openpyxl.styles import numbers
# from gui import ToolTip


def update_growth_before_retirement(wb, ws, username):
    contribution_401k = contribution_401k_entry.get()
    contribution_hsa = contribution_hsa_entry.get()
    contribution_stock = contribution_stock_entry.get()
    match_401 = match_401k_entry.get()
    match_hsa = match_hsa_entry.get()
    match_stock = match_stock_entry.get()

    ws['C75'] = int(contribution_401k)
    ws['D75'] = int(contribution_hsa)
    ws['E75'] = int(contribution_stock)
    ws['C76'] = int(match_401)
    ws['D76'] = int(match_hsa)
    ws['E76'] = int(match_stock)
    ws['C75'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['D75'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['E75'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C76'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['D76'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['E76'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip = None
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(self.tooltip, text=self.text, bg="yellow", relief=tk.SOLID, borderwidth=1)
        label.pack(ipadx=1)

    def leave(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()


def growth_before_retirement_window(growth_before_window, wb, ws, username):
    global contribution_401k_entry, contribution_hsa_entry, contribution_stock_entry, match_401k_entry, match_hsa_entry, match_stock_entry

    growth_before_window.geometry("300x300")
    contribution_401k_value = ws['C75'].value if ws['C75'].value else ""
    contribution_hsa_value = ws['D75'].value if ws['D75'].value else ""
    contribution_stock_value = ws['E75'].value if ws['E75'].value else ""
    match_401k_value = ws['C76'].value if ws['C75'].value else ""
    match_hsa_value = ws['D76'].value if ws['D76'].value else ""
    match_stock_value = ws['E76'].value if ws['E76'].value else ""

    contribution_401k_label = tk.Label(growth_before_window, text='Contribution 401k')
    contribution_401k_entry = tk.Entry(growth_before_window)
    contribution_401k_entry.insert(0, contribution_401k_value)
    contribution_401k_label.pack()
    contribution_401k_entry.pack()
    contribution_hsa_label = tk.Label(growth_before_window, text='Contribution HSA')
    contribution_hsa_entry = tk.Entry(growth_before_window)
    contribution_hsa_entry.insert(0, contribution_hsa_value)
    contribution_hsa_label.pack()
    contribution_hsa_entry.pack()
    contribution_stock_label = tk.Label(growth_before_window, text='Contribution Stock/Savings')
    contribution_stock_entry = tk.Entry(growth_before_window)
    contribution_stock_entry.insert(0, contribution_stock_value)
    contribution_stock_label.pack()
    contribution_stock_entry.pack()

    match_401k_label = tk.Label(growth_before_window, text='Match 401k')
    match_401k_entry = tk.Entry(growth_before_window)
    match_401k_entry.insert(0, match_401k_value)
    match_401k_label.pack()
    match_401k_entry.pack()
    match_hsa_label = tk.Label(growth_before_window, text='Match HSA')
    match_hsa_entry = tk.Entry(growth_before_window)
    match_hsa_entry.insert(0, match_hsa_value)
    match_hsa_label.pack()
    match_hsa_entry.pack()
    match_stock_label = tk.Label(growth_before_window, text='Match Stock/Savings')
    match_stock_entry = tk.Entry(growth_before_window)
    match_stock_entry.insert(0, match_stock_value)
    match_stock_label.pack()
    match_stock_entry.pack()

    ToolTip(contribution_401k_entry, "Enter your contribution to your 401k in USD")
    ToolTip(contribution_hsa_entry, "Enter your contribution to your HSA in USD")
    ToolTip(contribution_stock_entry, "Enter your contribution to your Stocks or Savings in USD")
    ToolTip(match_401k_entry, "Enter your match to your 401k in USD")
    ToolTip(match_hsa_entry, "Enter your match to your HSA in USD")
    ToolTip(match_stock_entry, "Enter your match to your Stocks or Savings in USD")

    submit_button = tk.Button(growth_before_window, text="Submit", command=lambda: update_growth_before_retirement(wb, ws, username))
    submit_button.pack()
