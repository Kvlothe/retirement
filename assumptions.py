import tkinter as tk
from openpyxl.styles import numbers


def update_assumptions(wb, ws, username):
    inflation = inflation_entry.get()
    cost_70 = cost_of_living_70_entry.get()
    cost_85 = cost_of_living_85_entry.get()
    social_cola = social_cola_entry.get()
    medicare_65 = medicare_65_entry.get()

    ws['C49'] = int(inflation)
    ws['C50'] = int(cost_70)
    ws['C51'] = int(cost_85)
    ws['C52'] = int(social_cola)
    ws['C53'] = int(medicare_65)
    ws['C49'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C50'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C51'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C52'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C53'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE


def assumptions_window(assumptions_input_window, wb, ws, username):
    global inflation_entry, cost_of_living_70_entry, cost_of_living_85_entry, social_cola_entry, medicare_65_entry

    assumptions_input_window.geometry("250x220")

    inflation_value = ws['C49'].value if ws['C49'].value else ""
    cost_70_value = ws['C50'].value if ws['C50'].value else ""
    cost_85_value = ws['C51'].value if ws['C51'].value else ""
    social_cola_value = ws['C52'].value if ws['C52'].value else ""
    medicare_65_value = ws['C53'].value if ws['C53'].value else ""

    inflation_label = tk.Label(assumptions_input_window, text="Inflation")
    inflation_label.pack()
    inflation_entry = tk.Entry(assumptions_input_window)
    inflation_entry.insert(0, inflation_value)
    inflation_entry.pack()
    cost_of_living_70_label = tk.Label(assumptions_input_window, text="Estimated Cost of living reduction at 70")
    cost_of_living_70_label.pack()
    cost_of_living_70_entry = tk.Entry(assumptions_input_window)
    cost_of_living_70_entry.insert(0, cost_70_value)
    cost_of_living_70_entry.pack()
    cost_of_living_85_label = tk.Label(assumptions_input_window, text="Estimated Cost of living increase at 85")
    cost_of_living_85_label.pack()
    cost_of_living_85_entry = tk.Entry(assumptions_input_window)
    cost_of_living_85_entry.insert(0, cost_85_value)
    cost_of_living_85_entry.pack()
    social_cola_label = tk.Label(assumptions_input_window, text="Social Security COLA, Estimated")
    social_cola_label.pack()
    social_cola_entry = tk.Entry(assumptions_input_window)
    social_cola_entry.insert(0, social_cola_value)
    social_cola_entry.pack()
    medicare_65_label = tk.Label(assumptions_input_window, text="Estimated Save with Medicare at 65")
    medicare_65_label.pack()
    medicare_65_entry = tk.Entry(assumptions_input_window)
    medicare_65_entry.insert(0, medicare_65_value)
    medicare_65_entry.pack()

    submit_button = tk.Button(assumptions_input_window, text="Submit", command=lambda: update_assumptions(wb, ws, username))
    submit_button.pack()

    assumptions_input_window.mainloop()
