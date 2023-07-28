import tkinter as tk
from openpyxl.styles import numbers


def update_current_profile(wb, ws, username):
    pre_tax = pre_tax_entry.get()
    after_tax = after_tax_entry.get()
    ws['C13'] = int(pre_tax)
    ws['C14'] = int(after_tax)
    ws['C13'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C14'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(f'{username}.xlsx')


def current_profile_window(current_window, wb, ws, username):
    global pre_tax_entry, after_tax_entry

    current_window.geometry("200x120")

    # Read in data from excel file
    pre_tax_value = ws['C13'].value if ws['C13'].value else ""
    after_tax_value = ws['C14'].value if ws['C14'].value else ""

    # Create Label and Entry for Current Portfolio
    pre_tax_label = tk.Label(current_window, text="Pre Tax - (IRA, 401K, 403B, HSA)")
    pre_tax_label.pack()
    pre_tax_entry = tk.Entry(current_window)
    pre_tax_entry.insert(0, pre_tax_value)  # populate the entry with value from excel
    pre_tax_entry.pack()

    after_tax_label = tk.Label(current_window, text="After Tax - (Stock, Cash, Savings)")
    after_tax_label.pack()
    after_tax_entry = tk.Entry(current_window)
    after_tax_entry.insert(0, after_tax_value)
    after_tax_entry.pack()

    submit_current = tk.Button(current_window, text="Submit", command=lambda: update_current_profile(wb, ws, username))
    submit_current.pack()

    current_window.mainloop()
