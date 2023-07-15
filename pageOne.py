import pandas as pd
import tkinter as tk
from openpyxl import load_workbook
from datetime import datetime

# Load the workbook.py
wb = load_workbook('sample.xlsx')

# Select the active worksheet
ws = wb.active


def update_excel():
    # Retrieve the input as a string and convert it to an integer
    year_input = int(year.get())
    # Place the input into the Excel worksheet
    ws['C5'] = year_input
    age_input = int(age.get())
    ws['C6'] = age_input
    retirement_age_input = int(retirement_age.get())
    ws['C7'] = retirement_age_input
    cell_c8_input = int(cell_c8.get())
    ws['C8'] = cell_c8_input
    cell_e8_input = int(cell_e8.get())
    ws['E8'] = cell_e8_input
    cell_c9_input = int(cell_c9.get())
    ws['C9'] = cell_c9_input
    cell_e9_input = int(cell_e9.get())
    ws['E9'] = cell_e9_input
    cell_c10_input = int(cell_c10.get())
    ws['C10'] = cell_c10_input
    cell_e10_input = int(cell_e10.get())
    ws['E10'] = cell_e10_input
    cell_c13_input = int(cell_c13.get())
    ws['C13'] = cell_c13_input
    cell_c14_input = int(cell_c14.get())
    ws['C14'] = cell_c14_input
    cell_c18_input = int(cell_c18.get())
    ws['C18'] = cell_c18_input
    cell_c19_input = int(cell_c19.get())
    ws['C19'] = cell_c19_input
    cell_c22_input = int(cell_c22.get())
    ws['C22'] = cell_c22_input
    cell_d22_input = int(cell_d22.get())
    ws['D22'] = cell_d22_input
    cell_e22_input = int(cell_e22.get())
    ws['E22'] = cell_e22_input
    cell_c23_input = int(cell_c23.get())
    ws['C23'] = cell_c23_input
    cell_d23_input = int(cell_d23.get())
    ws['D23'] = cell_d23_input
    cell_e23_input = int(cell_e23.get())
    ws['E23'] = cell_e23_input
    cell_c28_input = int(cell_c28.get())
    ws['C28'] = cell_c28_input
    cell_c29_input = int(cell_c29.get())
    ws['C29'] = cell_c29_input
    cell_c30_input = int(cell_c30.get())
    ws['C30'] = cell_c30_input
    cell_c31_input = int(cell_c31.get())
    ws['C31'] = cell_c31_input
    cell_c32_input = int(cell_c32.get())
    ws['C32'] = cell_c32_input
    cell_c33_input = int(cell_c33.get())
    ws['C33'] = cell_c33_input
    cell_c34_input = int(cell_c34.get())
    ws['C34'] = cell_c34_input
    cell_c35_input = int(cell_c35.get())
    ws['C35'] = cell_c35_input
    cell_c36_input = int(cell_c36.get())
    ws['C36'] = cell_c36_input
    cell_c37_input = int(cell_c37.get())
    ws['C37'] = cell_c37_input
    cell_c38_input = int(cell_c38.get())
    ws['C38'] = cell_c38_input

    # cell_c_input = int(cell_c.get())
    # ws['C'] = cell_c_input

    wb.save('sample.xlsx')


# Basic Info
window = tk.Tk()
basic_label = tk.Label(window, text="Basic Info")
year_label = tk.Label(window, text="Current Year:")
year = tk.Entry(window)
age_label = tk.Label(window, text="Age:")
age = tk.Entry(window)
retirement_age_label = tk.Label(text="Age to Retire:")
retirement_age = tk.Entry(window)
cell_c8_label = tk.Label(text="Year to take Social Security for spouse 1:")
cell_c8 = tk.Entry(window)
cell_e8_label = tk.Label(window, text="Year to take Social Security for spouse 2:")
cell_e8 = tk.Entry(window)
cell_c9_label = tk.Label(window, text="Year to take Pension for spouse 1:")
cell_c9 = tk.Entry(window)
cell_e9_label = tk.Label(window, text="Year to take Pension for spouse 2:")
cell_e9 = tk.Entry(window)
cell_c10_label = tk.Label(window, text="Year to Collect other income (annuity etc):")
cell_c10 = tk.Entry(window)
cell_e10_label = tk.Label(window, text="Other Income:")
cell_e10 = tk.Entry(window)

basic_label.grid(row=0, column=0)
year_label.grid(row=1, column=0)
year.grid(row=2, column=0)
age_label.grid(row=3, column=0)
age.grid(row=4, column=0)
retirement_age_label.grid(row=5, column=0)
retirement_age.grid(row=6, column=0)
cell_c8_label.grid(row=7, column=0)
cell_c8.grid(row=8, column=0)
cell_e8_label.grid(row=9, column=0)
cell_e8.grid(row=10, column=0)
cell_c9_label.grid(row=11, column=0)
cell_c9.grid(row=12, column=0)
cell_e9_label.grid(row=13, column=0)
cell_e9.grid(row=14, column=0)
cell_c10_label.grid(row=15, column=0)
cell_c10.grid(row=16, column=0)
cell_e10_label.grid(row=17, column=0)
cell_e10.grid(row=18, column=0)

# Current Portfolio
portfolio_label = tk.Label(window, text="Current Portfolio")
cell_c13_label = tk.Label(window, text="Pre-Tax:")
cell_c13 = tk.Entry(window)
cell_c14_label = tk.Label(window, text="After Tax:")
cell_c14 = tk.Entry(window)

portfolio_label.grid(row=0, column=2)
cell_c13_label.grid(row=1, column=2)
cell_c13.grid(row=2, column=2)
cell_c14_label.grid(row=3, column=2)
cell_c14.grid(row=4, column=2)

# SS at full retirement
ss_label = tk.Label(window, text="SS at Full Retirement")
cell_c18_label = tk.Label(window, text="Spouse 1:")
cell_c18 = tk.Entry(window)
cell_c19_label = tk.Label(window, text="Spouse 2:")
cell_c19 = tk.Entry(window)

ss_label.grid(row=7, column=2)
cell_c18_label.grid(row=8, column=2)
cell_c18.grid(row=9, column=2)
cell_c19_label.grid(row=10, column=2)
cell_c19.grid(row=11, column=2)

# Pension
pension_label = tk.Label(window, text="Pension")
cell_c22_label = tk.Label(window, text="Spouse 1:(63)")
cell_c22 = tk.Entry(window)
cell_d22_label = tk.Label(window, text="Spouse 1:(65)")
cell_d22 = tk.Entry(window)
cell_e22_label = tk.Label(window, text="Spouse 1:(lump)")
cell_e22 = tk.Entry(window)

pension_label.grid(row=13, column=2)
cell_c22_label.grid(row=14, column=2)
cell_c22.grid(row=15, column=2)
cell_d22_label.grid(row=16, column=2)
cell_d22.grid(row=17, column=2)
cell_e22_label.grid(row=18, column=2)
cell_e22.grid(row=19, column=2)

cell_c23_label = tk.Label(window, text="Spouse 2:(63)")
cell_c23 = tk.Entry(window)
cell_d23_label = tk.Label(window, text="Spouse 2:(65)")
cell_d23 = tk.Entry(window)
cell_e23_label = tk.Label(window, text="Spouse 2:(lump)")
cell_e23 = tk.Entry(window)

cell_c23_label.grid(row=20, column=2)
cell_c23.grid(row=21, column=2)
cell_d23_label.grid(row=22, column=2)
cell_d23.grid(row=23, column=2)
cell_e23_label.grid(row=24, column=2)
cell_e23.grid(row=25, column=2)

# Inputs
inputs_label = tk.Label(window, text="Inputs")
cell_c28_label = tk.Label(window, text="House payments(Monthly):")
cell_c28 = tk.Entry(window)
cell_c29_label = tk.Label(window, text="Insurance, Car, etc: Monthly")
cell_c29 = tk.Entry(window)
cell_c30_label = tk.Label(window, text="Gas: Monthly")
cell_c30 = tk.Entry(window)
cell_c31_label = tk.Label(window, text="Car payment: Monthly")
cell_c31 = tk.Entry(window)
cell_c32_label = tk.Label(window, text="utils(water, electric, phone, net: Monthly")
cell_c32 = tk.Entry(window)
cell_c33_label = tk.Label(window, text="Food/household:(Monthly)")
cell_c33 = tk.Entry(window)
cell_c34_label = tk.Label(window, text="Entertainment/Eat out:(Monthly)")
cell_c34 = tk.Entry(window)
cell_c35_label = tk.Label(window, text="Travel:(Monthly)")
cell_c35 = tk.Entry(window)
cell_c36_label = tk.Label(window, text="Misc and unplanned:(Monthly)")
cell_c36 = tk.Entry(window)
cell_c37_label = tk.Label(window, text="Hobbies: Monthly")
cell_c37 = tk.Entry(window)
cell_c38_label = tk.Label(window, text="Charity: Monthly")
cell_c38 = tk.Entry(window)

inputs_label.grid(row=0, column=4)
cell_c28_label.grid(row=1, column=4)
cell_c28.grid(row=2, column=4)
cell_c29_label.grid(row=3, column=4)
cell_c29.grid(row=4, column=4)
cell_c30_label.grid(row=5, column=4)
cell_c30.grid(row=6, column=4)
cell_c31_label.grid(row=7, column=4)
cell_c31.grid(row=8, column=4)
cell_c32_label.grid(row=9, column=4)
cell_c32.grid(row=10, column=4)
cell_c33_label.grid(row=11, column=4)
cell_c33.grid(row=12, column=4)
cell_c34_label.grid(row=13, column=4)
cell_c34.grid(row=14, column=4)
cell_c35_label.grid(row=15, column=4)
cell_c35.grid(row=16, column=4)
cell_c36_label.grid(row=17, column=4)
cell_c36.grid(row=18, column=4)
cell_c37_label.grid(row=19, column=4)
cell_c37.grid(row=20, column=4)
cell_c38_label.grid(row=21, column=4)
cell_c38.grid(row=22, column=4)

# cell__label = tk.Label(text=":")
# cell__label.grid()
# cell_ = tk.Entry()
# cell_.pack()

# Create a button that triggers the update_excel function when clicked
submit_button = tk.Button(window, text="Submit", command=update_excel)
submit_button.grid(row=30, column=4)


window.mainloop()
