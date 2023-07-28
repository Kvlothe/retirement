import tkinter as tk
from openpyxl.styles import numbers


def update_inputs(wb, ws, username):
    house_pay = house_payment_entry.get()
    insurance = insurance_entry.get()
    gas = gas_entry.get()
    car_pay = car_payment_entry.get()
    utilities = utilities_entry.get()
    food = food_household_entry.get()
    entertainment = entertainment_entry.get()
    travel = travel_entry.get()
    misc = misc_unplanned_entry.get()
    hobbies = hobbies_entry.get()
    charity = charity_entry.get()

    ws['C28'] = int(house_pay)
    ws['C29'] = int(insurance)
    ws['C30'] = int(gas)
    ws['C31'] = int(car_pay)
    ws['C32'] = int(utilities)
    ws['C33'] = int(food)
    ws['C34'] = int(entertainment)
    ws['C35'] = int(travel)
    ws['C36'] = int(misc)
    ws['C37'] = int(hobbies)
    ws['C38'] = int(charity)
    ws['C28'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C29'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C30'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C31'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C32'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C33'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C34'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C35'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C36'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C37'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C38'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(f'{username}.xlsx')


def inputs_window(input_window, wb, ws, username):
    global house_payment_entry, insurance_entry, gas_entry, car_payment_entry, utilities_entry, food_household_entry, entertainment_entry, travel_entry, misc_unplanned_entry, hobbies_entry, charity_entry

    input_window.geometry("400x600")

    house_payment_value = ws['C28'].value if ws['C28'].value else ""
    insurance_value = ws['C29'].value if ws['C29'].value else ""
    gas_value = ws['C30'].value if ws['C30'].value else ""
    car_payment_value = ws['C31'].value if ws['C31'].value else ""
    utilities_value = ws['C32'].value if ws['C32'].value else ""
    food_value = ws['C33'].value if ws['C33'].value else ""
    entertainment_value = ws['C34'].value if ws['C34'].value else""
    travel_value = ws['C35'].value if ws['C35'].value else ""
    misc_value = ws['C36'].value if ws['C36'].value else ""
    hobbies_value = ws['C37'].value if ws['C37'].value else ""
    charity_value = ws['C38'].value if ws['C38'].value else ""

    # Create Label and Entry for Expense estimates
    house_payment_label = tk.Label(input_window, text="House Payment")
    house_payment_label.pack()
    house_payment_entry = tk.Entry(input_window)
    house_payment_entry.insert(0, house_payment_value)
    house_payment_entry.pack()
    insurance_label = tk.Label(input_window, text="Insurance, Car, ETC")
    insurance_label.pack()
    insurance_entry = tk.Entry(input_window)
    insurance_entry.insert(0, insurance_value)
    insurance_entry.pack()
    gas_label = tk.Label(input_window, text="Gas")
    gas_label.pack()
    gas_entry = tk.Entry(input_window)
    gas_entry.insert(0, gas_value)
    gas_entry.pack()
    car_payment_label = tk.Label(input_window, text="Car Payment")
    car_payment_label.pack()
    car_payment_entry = tk.Entry(input_window)
    car_payment_entry.insert(0, car_payment_value)
    car_payment_entry.pack()
    utilities_label = tk.Label(input_window, text="Utilities (Water, Electric, Phone, Internet)")
    utilities_label.pack()
    utilities_entry = tk.Entry(input_window)
    utilities_entry.insert(0, utilities_value)
    utilities_entry.pack()
    food_household_label = tk.Label(input_window, text="Food and Household")
    food_household_label.pack()
    food_household_entry = tk.Entry(input_window)
    food_household_entry.insert(0, food_value)
    food_household_entry.pack()
    entertainment_label = tk.Label(input_window, text="Entertainment and Eating Out")
    entertainment_label.pack()
    entertainment_entry = tk.Entry(input_window)
    entertainment_entry.insert(0, entertainment_value)
    entertainment_entry.pack()
    travel_label = tk.Label(input_window, text="Travel")
    travel_label.pack()
    travel_entry = tk.Entry(input_window)
    travel_entry.insert(0, travel_value)
    travel_entry.pack()
    misc_unplanned_label = tk.Label(input_window, text="Miscellaneous and Unplanned")
    misc_unplanned_label.pack()
    misc_unplanned_entry = tk.Entry(input_window)
    misc_unplanned_entry.insert(0, misc_value)
    misc_unplanned_entry.pack()
    hobbies_label = tk.Label(input_window, text="Hobbies")
    hobbies_label.pack()
    hobbies_entry = tk.Entry(input_window)
    hobbies_entry.insert(0, hobbies_value)
    hobbies_entry.pack()
    charity_label = tk.Label(input_window, text="Charity")
    charity_label.pack()
    charity_entry = tk.Entry(input_window)
    charity_entry.insert(0, charity_value)
    charity_entry.pack()

    submit_basic = tk.Button(input_window, text="Submit", command=lambda: update_inputs(wb, ws, username))
    submit_basic.pack()

    input_window.mainloop()
