import tkinter as tk
from openpyxl.styles import numbers


def update_social_info(wb, ws, username):
    ss1 = ss_s1_entry.get()
    ss2 = ss_s2_entry.get()

    ws['C18'] = int(ss1)
    ws['C19'] = int(ss2)
    ws['C18'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws['C19'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(f'{username}.xlsx')


def social_security_window(security_window, wb, ws, username):
    global main_window, ss_s1_entry, ss_s2_entry

    security_window.geometry("200x120")

    ss_s1_value = ws['C18'].value if ws['C18'].value else ""
    ss_s2_value = ws['C19'].value if ws['C19'].value else ""

    # Create Label and Entry for Social Security
    ss_s1_label = tk.Label(security_window, text="Self")
    ss_s1_label.pack()
    ss_s1_entry = tk.Entry(security_window)
    ss_s1_entry.insert(0, ss_s1_value)
    ss_s1_entry.pack()
    ss_s2_label = tk.Label(security_window, text="Spouse")
    ss_s2_label.pack()
    ss_s2_entry = tk.Entry(security_window)
    ss_s2_entry.insert(0, ss_s2_value)
    ss_s2_entry.pack()

    submit_basic = tk.Button(security_window, text="Submit", command=lambda: update_social_info(wb, ws, username))
    submit_basic.pack()

    security_window.mainloop()
